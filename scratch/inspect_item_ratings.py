import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Try Hard\Desktop\Nexte\informes-andar\informes\Juan Pablo Herrera .xlsx", data_only=True)
sheet = wb['JULIO']

print("=== COLS 1-32 FOR ROWS 6, 7, 8, 9 ===")
for r in range(6, 10):
    row_str = f"Row {r:02d}:\n"
    for c in range(1, 33):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        fill = cell.fill
        color = fill.start_color.rgb if fill and fill.start_color and fill.start_color.rgb else "NONE"
        
        # We represent col name like A, B, C...
        col_letter = openpyxl.utils.get_column_letter(c)
        
        if val is not None or (color != "00000000" and color != "FFFFFFFF" and color != "NONE"):
            val_part = f"'{val}'" if val is not None else "None"
            row_str += f"  {col_letter}{r}: {val_part} (Color: {color})\n"
    print(row_str)
