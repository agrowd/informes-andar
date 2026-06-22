import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Try Hard\Desktop\Nexte\informes-andar\informes\Juan Pablo Herrera .xlsx", data_only=True)
sheet = wb['JULIO']

print("=== COMPLETE GRID ROWS 6-20 ===")
# Print column headers
col_headers = "Row | "
for c in range(1, 33):
    col_headers += f"{openpyxl.utils.get_column_letter(c):>3} "
print(col_headers)
print("-" * 140)

for r in range(6, 21):
    row_str = f"{r:03d} | "
    for c in range(1, 33):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        fill = cell.fill
        color = fill.start_color.rgb if fill and fill.start_color and fill.start_color.rgb else "NONE"
        
        # Determine cell display symbol
        if val is not None:
            # Show first letter of text or value
            symbol = str(val).strip()
            if len(symbol) > 3:
                symbol = symbol[:3]
            else:
                symbol = symbol.rjust(3)
        elif color == "FFA4C2F4":
            # Light blue color representing a check or option
            symbol = "BLU"
        elif color == "FFFF00FF":
            # Magenta color
            symbol = "MAG"
        elif color != "00000000" and color != "FFFFFFFF" and color != "NONE":
            # Other color
            symbol = color[:3]
        else:
            symbol = " . "
        
        row_str += f"{symbol:>3} "
    print(row_str)
