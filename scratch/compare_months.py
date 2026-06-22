import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Try Hard\Desktop\Nexte\informes-andar\informes\Juan Pablo Herrera .xlsx", data_only=True)

for sname in ['MAYO ', 'JULIO', 'AGOSTO', 'NOVIEMBRE']:
    sheet = wb[sname]
    # Check Item 1: Desarrollar fuerza (Cols A-D, Rows 8-10)
    # Check Item 4: Lanzar con brazo derecho (Cols M-P, Rows 8-10)
    # Check Item 9: Elongar miembros inferiores (Cols A-D, Rows 13-15)
    
    print(f"\nSheet {sname}:")
    # Desarrollar fuerza
    color_8 = sheet.cell(row=8, column=2).fill.start_color.rgb
    color_9 = sheet.cell(row=9, column=2).fill.start_color.rgb
    color_10 = sheet.cell(row=10, column=2).fill.start_color.rgb
    print(f"  Desarrollar fuerza (A-D) -> Row8: {color_8} | Row9: {color_9} | Row10: {color_10}")
    
    # Lanzar con brazo derecho
    color_l8 = sheet.cell(row=8, column=14).fill.start_color.rgb
    color_l9 = sheet.cell(row=9, column=14).fill.start_color.rgb
    color_l10 = sheet.cell(row=10, column=14).fill.start_color.rgb
    print(f"  Lanzar derecho (M-P)     -> Row8: {color_l8} | Row9: {color_l9} | Row10: {color_l10}")
    
    # Elongar miembros inferiores
    color_e13 = sheet.cell(row=13, column=2).fill.start_color.rgb
    color_e14 = sheet.cell(row=14, column=2).fill.start_color.rgb
    color_e15 = sheet.cell(row=15, column=2).fill.start_color.rgb
    print(f"  Elongar inf (A-D, Row11) -> Row13: {color_e13} | Row14: {color_e14} | Row15: {color_e15}")
